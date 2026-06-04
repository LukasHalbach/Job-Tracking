import os
import sqlite3
import hashlib
import json
import csv
from datetime import datetime, date
from flask import Flask, request, jsonify, render_template, send_file
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), "data", "timelog.db")

TASKS = [
    ("Roll Printing",               "Printing"),
    ("Sheet Printing",              "Printing"),
    ("Direct to Fabric Printing",   "Printing"),
    ("On-Site Printing",            "Printing"),
    ("Contour Cutting",             "Cutting"),
    ("Foam Cutting",                "Cutting"),
    ("Laminating",                  "Finishing"),
    ("Heat Press",                  "Finishing"),
    ("Vinyl Application",           "Finishing"),
    ("Table Work",                  "Finishing"),
    ("Fabric Finishing",            "Finishing"),
    ("Edge Welding",                "Finishing"),
    ("CNC Routing",                 "Fabrication"),
    ("Laser Engraving",             "Fabrication"),
    ("3D Printing",                 "Fabrication"),
    ("Metal Fabrication",           "Fabrication"),
    ("Wood Fabrication",            "Fabrication"),
    ("Embroidery",                  "Embroidery & Sewing"),
    ("Sewing",                      "Embroidery & Sewing"),
    ("On-Site Installation",        "Installation"),
    ("Vehicle Loading",             "Installation"),
    ("File Prep & Art Setup",       "Pre-Production"),
    ("Prepress",                    "Pre-Production"),
    ("Color Matching & Test Prints","Pre-Production"),
    ("Proof Creation & Output",     "Pre-Production"),
    ("Client Artwork Review",       "Pre-Production"),
    ("Morning Prep",                "Shop Operations"),
    ("Inventory Management",        "Shop Operations"),
    ("Ordering Material",           "Shop Operations"),
    ("Shipping",                    "Shop Operations"),
    ("Cleaning",                    "Shop Operations"),
    ("Machine Maintenance",         "Equipment Maintenance"),
    ("Vehicle Maintenance",         "Equipment Maintenance"),
    ("Loading Machines",            "Equipment Maintenance"),
    ("Admin",                       "Admin"),
    ("Client Communication",        "Admin"),
    ("Estimate Prep",               "Admin"),
    ("Vendor Communication",        "Admin"),
    ("Department Head Coordination","Admin"),
    ("Team Meeting",                "Admin"),
    ("Training",                    "Admin"),
    ("Assisting",                   "Admin"),
    ("Not Listed",                  "Other"),
]


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def hash_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                pin_hash TEXT NOT NULL,
                is_admin INTEGER DEFAULT 0,
                active INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_number TEXT NOT NULL,
                description TEXT DEFAULT '',
                active INTEGER DEFAULT 1,
                is_system INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                active INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS time_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                entry_date TEXT NOT NULL,
                job_id INTEGER,
                job_number TEXT NOT NULL,
                task_id INTEGER,
                task_name TEXT NOT NULL,
                category TEXT NOT NULL,
                hours REAL NOT NULL,
                description TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (employee_id) REFERENCES employees(id)
            );
        """)

        # Migrate: add is_system column to jobs if not present
        try:
            conn.execute("ALTER TABLE jobs ADD COLUMN is_system INTEGER DEFAULT 0")
        except Exception:
            pass

        # Seed the protected "Shop" system job
        shop = conn.execute("SELECT id FROM jobs WHERE job_number='Shop'").fetchone()
        if not shop:
            conn.execute(
                "INSERT INTO jobs (job_number, description, is_system) VALUES ('Shop', 'General shop work', 1)"
            )
        else:
            conn.execute("UPDATE jobs SET is_system=1 WHERE job_number='Shop'")

        # Seed the protected "Not Listed" system job
        not_listed = conn.execute("SELECT id FROM jobs WHERE job_number='Not Listed'").fetchone()
        if not not_listed:
            conn.execute(
                "INSERT INTO jobs (job_number, description, is_system) VALUES ('Not Listed', 'Invoice not yet in system', 1)"
            )
        else:
            conn.execute("UPDATE jobs SET is_system=1 WHERE job_number='Not Listed'")

        # Sync tasks — add new, reactivate returning, deactivate removed
        canonical = {name: category for name, category in TASKS}
        existing  = {row["name"]: row for row in
                     conn.execute("SELECT name, category, active FROM tasks").fetchall()}
        for name, category in TASKS:
            if name not in existing:
                conn.execute("INSERT INTO tasks (name, category) VALUES (?, ?)",
                             (name, category))
            else:
                conn.execute("UPDATE tasks SET category=?, active=1 WHERE name=?",
                             (category, name))
        for name in existing:
            if name not in canonical:
                conn.execute("UPDATE tasks SET active=0 WHERE name=?", (name,))

        # Roles tables
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS role_tasks (
                role_id INTEGER NOT NULL,
                task_id INTEGER NOT NULL,
                PRIMARY KEY (role_id, task_id),
                FOREIGN KEY (role_id) REFERENCES roles(id),
                FOREIGN KEY (task_id) REFERENCES tasks(id)
            );
        """)
        try:
            conn.execute("ALTER TABLE employees ADD COLUMN role_id INTEGER REFERENCES roles(id)")
        except Exception:
            pass

        # Seed a default admin if no employees exist
        count = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        if count == 0:
            conn.execute(
                "INSERT INTO employees (name, pin_hash, is_admin) VALUES (?, ?, 1)",
                ("Admin", hash_pin("0000")),
            )


# ── Auth helpers ─────────────────────────────────────────────────────────────

def verify_employee(employee_id: int, pin: str):
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM employees WHERE id=? AND active=1", (employee_id,)
        ).fetchone()
    if row and row["pin_hash"] == hash_pin(pin):
        return dict(row)
    return None


def verify_admin(employee_id: int, pin: str):
    emp = verify_employee(employee_id, pin)
    if emp and emp["is_admin"]:
        return emp
    return None


# ── API Routes ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/employees", methods=["GET"])
def list_employees():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, name, is_admin FROM employees WHERE active=1 ORDER BY name"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/auth", methods=["POST"])
def auth():
    data = request.json
    emp = verify_employee(data.get("employee_id"), str(data.get("pin", "")))
    if emp:
        allowed_task_ids = None
        if not emp["is_admin"] and emp.get("role_id"):
            with get_db() as conn:
                task_rows = conn.execute(
                    "SELECT task_id FROM role_tasks WHERE role_id=?", (emp["role_id"],)
                ).fetchall()
                not_listed = conn.execute(
                    "SELECT id FROM tasks WHERE name='Not Listed'"
                ).fetchone()
                allowed_task_ids = [r["task_id"] for r in task_rows]
                if not_listed and not_listed["id"] not in allowed_task_ids:
                    allowed_task_ids.append(not_listed["id"])
        return jsonify({
            "ok": True,
            "name": emp["name"],
            "is_admin": bool(emp["is_admin"]),
            "allowed_task_ids": allowed_task_ids,
        })
    return jsonify({"ok": False}), 401


@app.route("/api/jobs", methods=["GET"])
def list_jobs():
    show_all = request.args.get("all") == "1"
    with get_db() as conn:
        if show_all:
            rows = conn.execute(
                "SELECT * FROM jobs ORDER BY is_system DESC, active DESC, job_number"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM jobs WHERE active=1 ORDER BY is_system DESC, job_number"
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/jobs", methods=["POST"])
def add_job():
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    job_number = (data.get("job_number") or "").strip()
    if not job_number:
        return jsonify({"error": "Job number required"}), 400
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO jobs (job_number, description) VALUES (?, ?)",
            (job_number, data.get("description", "")),
        )
        return jsonify({"id": cur.lastrowid, "job_number": job_number}), 201


@app.route("/api/jobs/<int:job_id>", methods=["PUT"])
def update_job(job_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    with get_db() as conn:
        conn.execute(
            "UPDATE jobs SET job_number=?, description=?, active=? WHERE id=?",
            (data["job_number"], data.get("description", ""), int(data.get("active", 1)), job_id),
        )
    return jsonify({"ok": True})


@app.route("/api/jobs/<int:job_id>", methods=["DELETE"])
def delete_job(job_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized — admin PIN required"}), 403
    with get_db() as conn:
        job = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
        if not job:
            return jsonify({"error": "Job not found"}), 404
        if job["is_system"]:
            return jsonify({"error": "System jobs cannot be deleted"}), 403
        conn.execute("DELETE FROM jobs WHERE id=?", (job_id,))
    return jsonify({"ok": True})


@app.route("/api/jobs/import", methods=["POST"])
def import_jobs_csv():
    employee_id = request.form.get("employee_id")
    pin = request.form.get("pin", "")
    if not verify_admin(employee_id, pin):
        return jsonify({"error": "Unauthorized"}), 403

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    if not file.filename.lower().endswith(".csv"):
        return jsonify({"error": "File must be a .csv"}), 400

    stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline="")
    reader = csv.DictReader(stream)

    # Accept flexible column names
    headers = [h.strip().lower() for h in (reader.fieldnames or [])]
    number_col = next((h for h in reader.fieldnames or []
                       if h.strip().lower() in ("invoice", "invoice number",
                                                "job number", "job", "number")), None)
    desc_col   = next((h for h in reader.fieldnames or []
                       if h.strip().lower() in ("title", "description",
                                                "desc", "name")), None)

    if not number_col:
        return jsonify({"error": "Could not find an invoice/job number column. "
                        "Expected a column named: Invoice, Invoice Number, or Job Number"}), 400

    added = 0
    skipped = 0
    with get_db() as conn:
        for row in reader:
            number = (row.get(number_col) or "").strip()
            desc   = (row.get(desc_col)   or "").strip() if desc_col else ""
            if not number:
                skipped += 1
                continue
            existing = conn.execute(
                "SELECT id, is_system FROM jobs WHERE job_number=?", (number,)
            ).fetchone()
            if existing:
                skipped += 1
                continue
            # Never import a row that would shadow a system job name
            if number.lower() == 'shop':
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO jobs (job_number, description) VALUES (?, ?)",
                (number, desc),
            )
            added += 1

    return jsonify({"added": added, "skipped": skipped})


@app.route("/api/tasks", methods=["GET"])
def list_tasks():
    show_all = request.args.get("all") == "1"
    with get_db() as conn:
        if show_all:
            rows = conn.execute(
                "SELECT * FROM tasks ORDER BY active DESC, name"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM tasks WHERE active=1 ORDER BY name"
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/tasks", methods=["POST"])
def add_task():
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    name = (data.get("name") or "").strip()
    category = (data.get("category") or "").strip()
    if not name or not category:
        return jsonify({"error": "Name and category required"}), 400
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO tasks (name, category) VALUES (?, ?)", (name, category)
        )
        return jsonify({"id": cur.lastrowid, "name": name, "category": category}), 201


@app.route("/api/tasks/<int:task_id>", methods=["PUT"])
def update_task(task_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    with get_db() as conn:
        conn.execute(
            "UPDATE tasks SET name=?, category=?, active=? WHERE id=?",
            (data["name"], data["category"], int(data.get("active", 1)), task_id),
        )
    return jsonify({"ok": True})


@app.route("/api/tasks/<int:task_id>", methods=["DELETE"])
def delete_task(task_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized — admin PIN required"}), 403
    with get_db() as conn:
        task = conn.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
        if not task:
            return jsonify({"error": "Task not found"}), 404
        if task["name"] == "Not Listed":
            return jsonify({"error": "'Not Listed' is a required system task and cannot be deleted"}), 403
        conn.execute("DELETE FROM tasks WHERE id=?", (task_id,))
    return jsonify({"ok": True})


@app.route("/api/entries", methods=["GET"])
def list_entries():
    employee_id = request.args.get("employee_id")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    all_employees = request.args.get("all_employees") == "1"

    query = """
        SELECT te.*, e.name as employee_name, j.description as job_description
        FROM time_entries te
        JOIN employees e ON te.employee_id = e.id
        LEFT JOIN jobs j ON te.job_id = j.id
        WHERE 1=1
    """
    params = []

    if not all_employees and employee_id:
        query += " AND te.employee_id = ?"
        params.append(employee_id)

    if date_from:
        query += " AND te.entry_date >= ?"
        params.append(date_from)

    if date_to:
        query += " AND te.entry_date <= ?"
        params.append(date_to)

    query += " ORDER BY te.entry_date DESC, te.created_at DESC LIMIT 200"

    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/entries", methods=["POST"])
def add_entry():
    data = request.json
    emp = verify_employee(data.get("employee_id"), str(data.get("pin", "")))
    if not emp:
        return jsonify({"error": "Unauthorized"}), 401

    job_number = (data.get("job_number") or "").strip()
    task_name = (data.get("task_name") or "").strip()
    category = (data.get("category") or "").strip()
    hours = float(data.get("hours") or 0)

    if not job_number:
        return jsonify({"error": "Job number required"}), 400
    if not task_name:
        return jsonify({"error": "Task required"}), 400
    if hours <= 0:
        return jsonify({"error": "Hours must be greater than 0"}), 400

    entry_date = data.get("entry_date") or date.today().isoformat()

    with get_db() as conn:
        # Job must exist in the active jobs list
        job_row = conn.execute(
            "SELECT id FROM jobs WHERE job_number=? AND active=1", (job_number,)
        ).fetchone()
        if not job_row:
            return jsonify({"error": f"'{job_number}' is not on the job list. Please select a job from the dropdown."}), 400
        job_id = job_row["id"]

        # Resolve task_id
        task_row = conn.execute(
            "SELECT id FROM tasks WHERE name=? AND active=1", (task_name,)
        ).fetchone()
        task_id = task_row["id"] if task_row else None

        cur = conn.execute(
            """INSERT INTO time_entries
               (employee_id, entry_date, job_id, job_number, task_id, task_name,
                category, hours, description, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                emp["id"], entry_date, job_id, job_number, task_id, task_name,
                category, hours,
                data.get("description", ""),
                data.get("notes", ""),
            ),
        )
        return jsonify({"id": cur.lastrowid}), 201


@app.route("/api/entries/<int:entry_id>", methods=["PUT"])
def update_entry(entry_id):
    data = request.json
    emp = verify_employee(data.get("employee_id"), str(data.get("pin", "")))
    if not emp:
        return jsonify({"error": "Unauthorized"}), 401

    with get_db() as conn:
        existing = conn.execute(
            "SELECT * FROM time_entries WHERE id=?", (entry_id,)
        ).fetchone()
        if not existing:
            return jsonify({"error": "Entry not found"}), 404
        # Only admin or the owning employee can edit
        if not emp["is_admin"] and existing["employee_id"] != emp["id"]:
            return jsonify({"error": "Forbidden"}), 403

        job_number = (data.get("job_number") or "").strip()
        task_name = (data.get("task_name") or "").strip()
        hours = float(data.get("hours") or 0)

        job_row = conn.execute(
            "SELECT id FROM jobs WHERE job_number=? AND active=1", (job_number,)
        ).fetchone()
        if not job_row:
            return jsonify({"error": f"'{job_number}' is not on the job list. Please select a job from the dropdown."}), 400
        job_id = job_row["id"]

        task_row = conn.execute(
            "SELECT id FROM tasks WHERE name=?", (task_name,)
        ).fetchone()
        task_id = task_row["id"] if task_row else None

        conn.execute(
            """UPDATE time_entries SET
               entry_date=?, job_id=?, job_number=?, task_id=?, task_name=?,
               category=?, hours=?, description=?, notes=?,
               updated_at=datetime('now')
               WHERE id=?""",
            (
                data.get("entry_date", existing["entry_date"]),
                job_id, job_number, task_id, task_name,
                data.get("category", existing["category"]),
                hours,
                data.get("description", existing["description"]),
                data.get("notes", existing["notes"]),
                entry_id,
            ),
        )
    return jsonify({"ok": True})


@app.route("/api/entries/<int:entry_id>", methods=["DELETE"])
def delete_entry(entry_id):
    data = request.json
    emp = verify_employee(data.get("employee_id"), str(data.get("pin", "")))
    if not emp:
        return jsonify({"error": "Unauthorized"}), 401

    with get_db() as conn:
        existing = conn.execute(
            "SELECT * FROM time_entries WHERE id=?", (entry_id,)
        ).fetchone()
        if not existing:
            return jsonify({"error": "Entry not found"}), 404
        if not emp["is_admin"] and existing["employee_id"] != emp["id"]:
            return jsonify({"error": "Forbidden"}), 403
        conn.execute("DELETE FROM time_entries WHERE id=?", (entry_id,))
    return jsonify({"ok": True})


# ── Admin: Employees ──────────────────────────────────────────────────────────

@app.route("/api/admin/employees", methods=["GET"])
def admin_list_employees():
    data = request.args
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    with get_db() as conn:
        rows = conn.execute(
            """SELECT e.id, e.name, e.is_admin, e.active, e.role_id,
                      r.name as role_name
               FROM employees e
               LEFT JOIN roles r ON e.role_id = r.id
               ORDER BY e.name"""
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/employees", methods=["POST"])
def admin_add_employee():
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    name = (data.get("name") or "").strip()
    pin = str(data.get("new_pin") or "").strip()
    if not name or not pin:
        return jsonify({"error": "Name and PIN required"}), 400
    role_id = data.get("role_id") or None
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO employees (name, pin_hash, is_admin, role_id) VALUES (?, ?, ?, ?)",
            (name, hash_pin(pin), int(data.get("is_admin", 0)), role_id),
        )
    return jsonify({"id": cur.lastrowid, "name": name}), 201


@app.route("/api/admin/employees/<int:emp_id>", methods=["PUT"])
def admin_update_employee(emp_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    role_id = data.get("role_id") or None
    with get_db() as conn:
        if data.get("new_pin"):
            conn.execute(
                "UPDATE employees SET name=?, pin_hash=?, is_admin=?, active=?, role_id=? WHERE id=?",
                (data["name"], hash_pin(str(data["new_pin"])),
                 int(data.get("is_admin", 0)), int(data.get("active", 1)), role_id, emp_id),
            )
        else:
            conn.execute(
                "UPDATE employees SET name=?, is_admin=?, active=?, role_id=? WHERE id=?",
                (data["name"], int(data.get("is_admin", 0)),
                 int(data.get("active", 1)), role_id, emp_id),
            )
    return jsonify({"ok": True})


@app.route("/api/employees/<int:emp_id>/change-pin", methods=["POST"])
def change_pin(emp_id):
    data = request.json
    old_pin = str(data.get("old_pin", ""))
    new_pin = str(data.get("new_pin", ""))
    if not verify_employee(emp_id, old_pin):
        return jsonify({"error": "Current PIN is incorrect"}), 401
    if not new_pin:
        return jsonify({"error": "New PIN is required"}), 400
    with get_db() as conn:
        conn.execute("UPDATE employees SET pin_hash=? WHERE id=?",
                     (hash_pin(new_pin), emp_id))
    return jsonify({"ok": True})


# ── Roles ────────────────────────────────────────────────────────────────────

@app.route("/api/roles", methods=["GET"])
def list_roles():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM roles ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/roles", methods=["POST"])
def add_role():
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Role name required"}), 400
    with get_db() as conn:
        cur = conn.execute("INSERT INTO roles (name) VALUES (?)", (name,))
        return jsonify({"id": cur.lastrowid, "name": name}), 201


@app.route("/api/roles/<int:role_id>", methods=["PUT"])
def update_role(role_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Role name required"}), 400
    with get_db() as conn:
        conn.execute("UPDATE roles SET name=? WHERE id=?", (name, role_id))
    return jsonify({"ok": True})


@app.route("/api/roles/<int:role_id>", methods=["DELETE"])
def delete_role(role_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized — admin PIN required"}), 403
    with get_db() as conn:
        conn.execute("UPDATE employees SET role_id=NULL WHERE role_id=?", (role_id,))
        conn.execute("DELETE FROM role_tasks WHERE role_id=?", (role_id,))
        conn.execute("DELETE FROM roles WHERE id=?", (role_id,))
    return jsonify({"ok": True})


@app.route("/api/roles/<int:role_id>/tasks", methods=["GET"])
def get_role_tasks(role_id):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT task_id FROM role_tasks WHERE role_id=?", (role_id,)
        ).fetchall()
    return jsonify([r["task_id"] for r in rows])


@app.route("/api/roles/<int:role_id>/tasks", methods=["PUT"])
def set_role_tasks(role_id):
    data = request.json
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403
    task_ids = data.get("task_ids", [])
    with get_db() as conn:
        conn.execute("DELETE FROM role_tasks WHERE role_id=?", (role_id,))
        for tid in task_ids:
            conn.execute(
                "INSERT OR IGNORE INTO role_tasks (role_id, task_id) VALUES (?, ?)",
                (role_id, int(tid)),
            )
    return jsonify({"ok": True})


# ── Admin: Flagged Entries ────────────────────────────────────────────────────

@app.route("/api/admin/flagged-entries", methods=["GET"])
def admin_flagged_entries():
    data = request.args
    if not verify_admin(data.get("employee_id"), str(data.get("pin", ""))):
        return jsonify({"error": "Unauthorized"}), 403

    with get_db() as conn:
        # Entries whose job_number no longer exists as an active job,
        # OR whose task_name no longer exists in the system at all.
        # Deactivated jobs/tasks do NOT flag old entries — only hard-deleted ones do.
        # "Not Listed" job/task are intentional placeholders — excluded.
        rows = conn.execute("""
            SELECT te.id, te.entry_date, te.job_number, te.task_name,
                   te.category, te.hours, te.description, te.notes,
                   e.name AS employee_name,
                   CASE WHEN j.id IS NULL THEN 1 ELSE 0 END AS bad_job,
                   CASE WHEN t.id IS NULL THEN 1 ELSE 0 END AS bad_task
            FROM time_entries te
            JOIN employees e ON te.employee_id = e.id
            LEFT JOIN jobs j
                   ON j.job_number = te.job_number
            LEFT JOIN tasks t
                   ON t.name = te.task_name
            WHERE te.job_number != 'Not Listed'
              AND te.task_name  != 'Not Listed'
              AND (j.id IS NULL OR t.id IS NULL)
            ORDER BY te.entry_date DESC, te.id DESC
            LIMIT 500
        """).fetchall()

        active_jobs  = conn.execute(
            "SELECT job_number FROM jobs WHERE active=1 ORDER BY job_number"
        ).fetchall()
        active_tasks = conn.execute(
            "SELECT name, category FROM tasks WHERE active=1 ORDER BY name"
        ).fetchall()

    return jsonify({
        "entries": [dict(r) for r in rows],
        "active_jobs":  [r["job_number"] for r in active_jobs],
        "active_tasks": [{"name": r["name"], "category": r["category"]} for r in active_tasks],
    })


# ── Export ────────────────────────────────────────────────────────────────────

@app.route("/api/export", methods=["GET"])
def export_excel():
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    employee_id = request.args.get("employee_id", "")
    emp_pin = request.args.get("pin", "")

    # Must be authenticated
    if not verify_employee(int(employee_id) if employee_id else 0, emp_pin):
        return jsonify({"error": "Unauthorized"}), 401

    query = """
        SELECT te.entry_date, e.name as employee_name, te.job_number,
               te.task_name, te.category, te.hours, te.description, te.notes,
               te.created_at, te.updated_at
        FROM time_entries te
        JOIN employees e ON te.employee_id = e.id
        WHERE 1=1
    """
    params = []
    if date_from:
        query += " AND te.entry_date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND te.entry_date <= ?"
        params.append(date_to)
    query += " ORDER BY te.entry_date, e.name, te.created_at"

    with get_db() as conn:
        rows = conn.execute(query, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time Entries"

    header_fill = PatternFill("solid", fgColor="1E5C1E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Date", "Employee", "Job / Invoice #", "Task", "Category",
               "Hours", "Description", "Notes"]
    col_widths = [12, 20, 18, 30, 25, 8, 35, 35]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            row["entry_date"],
            row["employee_name"],
            row["job_number"],
            row["task_name"],
            row["category"],
            row["hours"],
            row["description"],
            row["notes"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = wrap
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(len(rows) + 1, 2)}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary by Employee")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18

    s_headers = ["Employee", "Total Hours", "Date From", "Date To"]
    for col, h in enumerate(s_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Aggregate by employee
    from collections import defaultdict
    emp_hours = defaultdict(float)
    emp_dates = defaultdict(list)
    for row in rows:
        emp_hours[row["employee_name"]] += row["hours"]
        emp_dates[row["employee_name"]].append(row["entry_date"])

    for r, (emp_name, total) in enumerate(sorted(emp_hours.items()), 2):
        dates = emp_dates[emp_name]
        ws2.cell(row=r, column=1, value=emp_name)
        ws2.cell(row=r, column=2, value=round(total, 2))
        ws2.cell(row=r, column=3, value=min(dates))
        ws2.cell(row=r, column=4, value=max(dates))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"time_entries_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=False)
